# Automation w/ Python, Excel Spreadsheets- automatically process thousands of spreadsheets in under 1s
# Change price, add chart

# Imports
import openpyxl as xl
from openpyxl.chart import BarChart, Reference  # Import 2 classes
from openpyxl.utils import get_column_letter  # For iter_rows, minimize memory usage for large datasets
from openpyxl.styles import NamedStyle
import os, random

from multiprocessing import Pool  # Parallelism

# Process
def process_workbook(filename, sheet_names, discount, output):  # Directory
    '''
    Processes given sheets in an Excel workbook, adds discount, saves result
    
    :param filename: Excel file path
    :param sheet_names: sheet names list
    :param discount: discount
    :param output: output Excel file path
    '''
    
    if not os.path.exists(filename):
        raise FileNotFoundError(f'{filename} does not exist')
    
    try:
        wb = xl.load_workbook(filename)
        
        if 'dollar_style' not in wb.named_styles:
            dollar_unit = NamedStyle(name='dollar_style', number_format='"$"#,##0.00')
            wb.add_named_style(dollar_unit)  # Ensures style is added to wb
        
        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                print(f'Sheet "{sheet_name}" does not exist. Skipping...')
                continue  # Skips missing sheets
            sheet = wb[sheet_name]
            print(f'Processing File "{filename}" Sheet "{sheet_name}"...')

            # cell = sheet['a1']
            # cell = sheet.cell(1, 1)  # Alternative
            # print(cell.value)
            # print(sheet.max_row)

            # for i in range(2, sheet.max_row + 1):
                # cell = sheet.cell(i, 3)  # Row i, col 3
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=3):
                cell = row[0]
                if isinstance(cell.value, (int, float)):
                    # corrected_price = f'${round(cell.value * (1 - discount), 2)}'  # Adds 10% discount, round to 2 decimal, add $
                    # corrected_price_cell = sheet.cell(i, 5)  
                    # corrected_price_cell.value = corrected_pricec
                    
                    corrected_price = round(cell.value * (1 - discount), 2)
                    # Set and put corrected price on col 5
                    output_col = get_column_letter(5)  # col e
                    sheet[f'{output_col}{cell.row}'] = corrected_price
                    
                    # Apply dollar unit
                    price_cell = sheet[f'{output_col}{cell.row}']
                    price_cell.style = 'dollar_style'  # Add 'dollar_style' directly instead...

            # Chart
            values = Reference(sheet, 
                    min_row=2, 
                    max_row=sheet.max_row, 
                    min_col=5, 
                    max_col=5
            )

            chart = BarChart()
            chart.title = f'Corrected Prices ({sheet_name})'
            chart.x_axis.title = 'Items'
            chart.y_axis.title = 'Price'
            # chart.style = 2
            chart.add_data(values)  # titles_from_data=True
            
            col_letter = get_column_letter(sheet.max_column + 2)
            sheet.add_chart(chart, f'{col_letter}2')  # Add chart...

        wb.save(output)
        wb.close()
        print(f'Processing complete. File saved as "{output}"')
    
    except Exception as e:
        print(f'Error: {e}')

# Parallel processing Wrapper
def process_file(args):
    filename, sheet_names, discount, output = args
    output_file = os.path.join(output, os.path.basename(filename).replace('.xlsx', '_processed.xlsx'))
    process_workbook(filename, sheet_names, discount, output_file)

# Processes multiple Excel files in parallel
def parallel_process(filenames, sheet_names, discount, output_dir):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)  # Ensures output directory exists
    
    # Prepares args
    args = [(filename, sheet_names, discount, output_dir) for filename in filenames]
    
    # Multiprocessing, concurrency
    with Pool() as pool:
        pool.map(process_file, args)
    
    print(f'All files processed and saved to "{output_dir}"')

# Get Excel files in directory
def get_excel_files(directory):
    return [f for f in os.listdir(directory) if f.endswith('.xlsx')]

# Test Excel file generator
def generate_excel(transaction_id_start, product_id_range, price_range, num_sheets):
    '''
    :param transaction_id_start: int
    :param product_id_range: [int, int]
    :param price_range: [float, float]
    '''
    wb = xl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    # ws = wb.active
    dollar_unit = NamedStyle(name='dollar_style', number_format='"$"#,##0.00')
    
    for sheet in range(1, num_sheets + 1):
        # Create new sheet per iteration
        if sheet >= 1:
            wb.create_sheet(title=f'Sheet{sheet}')
        ws = wb[f'Sheet{sheet}']
        
        ws.append(['transaction_id', 'product_id', 'price'])  # Header

        for i in range(1, 101):  # Fill rows
            transaction_id_start += 1
            product_id = random.randint(product_id_range[0], product_id_range[1])
            # price = f'${round(random.uniform(price_range[0], price_range[1]), 2)}'
            price = round(random.uniform(price_range[0], price_range[1]), 2)
            ws.append([transaction_id_start, product_id, price])

            # Apply dollar unit
            price_cell = ws.cell(row=i+1, column=3)
            price_cell.style = dollar_unit
    
    wb.save('transactions_generated.xlsx')
    print('Test Excel generated and saved as transactions_generated.xlsx')


# Main
if __name__ == '__main__':
    # Test Excel file generator
    transaction_id_start = 2000
    product_id_range = [1, 100]
    price_range = [5.0, 50.0]
    num_sheets = 3
    
    # generate_excel(transaction_id_start, product_id_range, price_range, num_sheets)  # Comment out to stop generating
    
    # filenames_dir = './excel files/'
    # filenames = get_excel_files(filenames_dir)  # Automatically gets Excel files in directory
    
    # Process
    filenames = ['transactions_generated.xlsx', 'transactions.xlsx']
    sheet_names = ['Sheet1', 'Sheet2', 'Sheet3']
    discount = 0.1  # 10% discount
    output_dir = 'processed_files'  # Output directory
    
    parallel_process(filenames, sheet_names, discount, output_dir)  # Process Excel files
    # Can also add legend, colors, other charts...
    # Rounded to 2 decimal